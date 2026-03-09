import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import google.generativeai as genai
import threading
import time

class ExcelTranslatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 翻譯工具 (Gemini API)")
        self.root.geometry("550x450")
        
        self.filepath = None
        
        # API Key 區塊
        tk.Label(root, text="請輸入 Gemini API Key:").pack(pady=(15, 5))
        self.api_key_entry = tk.Entry(root, width=60, show="*")
        self.api_key_entry.pack(pady=5)
        
        # 匯入區塊
        self.import_btn = tk.Button(root, text="1. 點選匯入 Excel 檔案", command=self.import_file, width=30, height=2)
        self.import_btn.pack(pady=15)
        self.file_label = tk.Label(root, text="尚未選擇檔案", fg="gray")
        self.file_label.pack(pady=5)
        
        # 狀態與日誌
        self.log_text = tk.Text(root, height=10, width=65, state=tk.DISABLED)
        self.log_text.pack(pady=10)
        
        # 匯出區塊
        self.export_btn = tk.Button(root, text="2. 開始翻譯並選擇匯出路徑", command=self.start_processing, width=30, height=2, bg="lightblue")
        self.export_btn.pack(pady=10)

    def log(self, msg):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def import_file(self):
        self.filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if self.filepath:
            self.file_label.config(text=self.filepath, fg="black")
            self.log(f"已匯入檔案: {self.filepath}")

    def start_processing(self):
        if not self.filepath:
            messagebox.showwarning("提示", "請先選擇 Excel 檔案！")
            return
        
        api_key = self.api_key_entry.get().strip()
        if not api_key:
            messagebox.showwarning("提示", "請輸入 Gemini API Key！")
            return
            
        export_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="選擇匯出路徑"
        )
        
        if not export_path:
            return

        self.export_btn.config(state=tk.DISABLED)
        self.import_btn.config(state=tk.DISABLED)
        # 使用執行緒處理，避免畫面卡死
        threading.Thread(target=self.process_excel, args=(api_key, export_path), daemon=True).start()

    def translate_text(self, model, text):
        if text is None or str(text).strip() == "":
            return ""
        try:
            prompt = f"請將以下文字(可能是日文或英文)翻譯成繁體中文。請直接回覆翻譯結果，不要加上引號、註解或原文：\n{text}"
            response = model.generate_content(prompt)
            return response.text.strip()
        except Exception as e:
            return f"[翻譯失敗: {e}]"

    def process_excel(self, api_key, export_path):
        try:
            genai.configure(api_key=api_key)
            # 使用 flash 模型，速度快且適合文本翻譯
            model = genai.GenerativeModel('gemini-2.5-flash') 
            
            self.log("正在載入 Excel 檔案...")
            wb = openpyxl.load_workbook(self.filepath)
            ws = wb.active
            
            self.log("正在重新編排欄位...")
            # 為了避免插入時影響後續欄位的索引，必須「由右至左」插入
            # 原本欄位為 A(1), B(2), C(3), D(4)
            # 在 D 右側插入 (相當於在第5欄前插入)
            ws.insert_cols(5)
            # 在 C 右側插入 (相當於在第4欄前插入)
            ws.insert_cols(4)
            # 在 B 右側插入 (相當於在第3欄前插入)
            ws.insert_cols(3)
            
            # 經過上述插入後，原欄位對應如下：
            # 原 B (現第2欄) -> 翻譯後放入 C (現第3欄)
            # 原 C (現第4欄) -> 翻譯後放入 E (現第5欄)
            # 原 D (現第6欄) -> 翻譯後放入 G (現第7欄)
            
            max_row = ws.max_row
            self.log(f"共發現 {max_row} 列資料，開始呼叫 Gemini API 翻譯...")
            
            # 從第一列開始處理
            for row in range(1, max_row + 1):
                cell_b = ws.cell(row=row, column=2).value
                cell_d = ws.cell(row=row, column=4).value
                cell_f = ws.cell(row=row, column=6).value
                
                # 若儲存格有內容才進行翻譯
                if cell_b:
                    ws.cell(row=row, column=3).value = self.translate_text(model, cell_b)
                if cell_d:
                    ws.cell(row=row, column=5).value = self.translate_text(model, cell_d)
                if cell_f:
                    ws.cell(row=row, column=7).value = self.translate_text(model, cell_f)
                    
                self.log(f"完成第 {row}/{max_row} 列...")
                
                # 稍微暫停以避免觸發 Gemini API 的請求次數限制 (Rate limit)
                time.sleep(1.5)
                
            self.log("正在儲存檔案...")
            wb.save(export_path)
            self.log("處理完成！")
            messagebox.showinfo("成功", f"檔案已成功翻譯並匯出至：\n{export_path}")
            
        except Exception as e:
            self.log(f"發生錯誤: {e}")
            messagebox.showerror("錯誤", f"處理時發生不可預期的錯誤：\n{e}")
        finally:
            self.export_btn.config(state=tk.NORMAL)
            self.import_btn.config(state=tk.NORMAL)

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelTranslatorApp(root)
    root.mainloop()